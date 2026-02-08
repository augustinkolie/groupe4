import os
from datetime import datetime
from django.conf import settings
from django.db.models import Avg, Max, Min, Count
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from monitoring.models import Station, Reading
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class ReportGenerator:
    def __init__(self, stations, start_date, end_date, report_type='CUSTOM'):
        self.stations = stations
        self.start_date = start_date
        self.end_date = end_date
        self.report_type = report_type
        self.styles = getSampleStyleSheet()
        
    def generate_pdf(self, filename):
        """Génère un rapport PDF professionnel"""
        # Utiliser un buffer au lieu d'un fichier
        buffer = io.BytesIO()
        
        # Créer le document PDF
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#10b981'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        title = Paragraph(f"Rapport de Qualité de l'Air EcoWatch", title_style)
        story.append(title)
        
        # Informations du rapport
        info_style = ParagraphStyle(
            'Info',
            parent=self.styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#6b7280'),
            alignment=TA_CENTER
        )
        
        period = Paragraph(f"Période: {self.start_date.strftime('%d/%m/%Y')} - {self.end_date.strftime('%d/%m/%Y')}", info_style)
        generated = Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", info_style)
        stations_count = Paragraph(f"Stations: {self.stations.count()}", info_style)
        
        story.append(period)
        story.append(generated)
        story.append(stations_count)
        story.append(Spacer(1, 30))
        
        # Message d'information
        info_para = Paragraph(
            "<b>Résumé:</b> Ce rapport présente les données de qualité de l'air pour les stations sélectionnées sur la période spécifiée.",
            self.styles['Normal']
        )
        story.append(info_para)
        story.append(Spacer(1, 20))
        
        # Vérifier s'il y a des données
        total_readings = 0
        
        # Pour chaque station
        for station in self.stations:
            readings = Reading.objects.filter(
                station=station,
                timestamp__date__gte=self.start_date,
                timestamp__date__lte=self.end_date
            )
            
            if not readings.exists():
                # Ajouter un message pour cette station
                station_title = Paragraph(f"<b>{station.name}</b>", self.styles['Heading2'])
                story.append(station_title)
                no_data_msg = Paragraph(
                    f"<i>Aucune donnée disponible pour cette période</i>",
                    self.styles['Normal']
                )
                story.append(no_data_msg)
                story.append(Spacer(1, 20))
                continue
            
            total_readings += readings.count()
            
            # Titre de la station
            station_title = Paragraph(f"<b>{station.name}</b>", self.styles['Heading2'])
            story.append(station_title)
            story.append(Spacer(1, 12))
            
            # Statistiques
            stats = readings.aggregate(
                avg_iqa=Avg('iqa'),
                max_iqa=Max('iqa'),
                min_iqa=Min('iqa'),
                avg_pm25=Avg('pm25'),
                avg_co=Avg('co'),
                count=Count('id')
            )
            
            stats_data = [
                ['Indicateur', 'Valeur'],
                ['IQA Moyen', f"{stats['avg_iqa']:.1f}" if stats['avg_iqa'] else 'N/A'],
                ['IQA Maximum', f"{stats['max_iqa']:.1f}" if stats['max_iqa'] else 'N/A'],
                ['IQA Minimum', f"{stats['min_iqa']:.1f}" if stats['min_iqa'] else 'N/A'],
                ['PM2.5 Moyen', f"{stats['avg_pm25']:.1f} µg/m³" if stats['avg_pm25'] else 'N/A'],
                ['CO Moyen', f"{stats['avg_co']:.1f} mg/m³" if stats['avg_co'] else 'N/A'],
                ['Nombre de relevés', str(stats['count'])],
            ]
            
            stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
            ]))
            
            story.append(stats_table)
            story.append(Spacer(1, 20))
            
            # Graphique d'évolution IQA
            chart_buffer = self._generate_iqa_chart(readings, station.name)
            if chart_buffer:
                img = Image(chart_buffer, width=5*inch, height=3*inch)
                story.append(img)
                story.append(Spacer(1, 20))
            
            story.append(PageBreak())
        
        # Construire le PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def _generate_iqa_chart(self, readings, station_name):
        """Génère un graphique d'évolution de l'IQA"""
        try:
            data = readings.order_by('timestamp').values_list('timestamp', 'iqa')
            if not data:
                return None
            
            timestamps = [d[0] for d in data]
            iqas = [d[1] if d[1] else 0 for d in data]
            
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.plot(timestamps, iqas, linewidth=2, color='#10b981', marker='o', markersize=4)
            ax.fill_between(timestamps, iqas, alpha=0.3, color='#10b981')
            
            # Zones de couleur
            ax.axhspan(0, 50, alpha=0.1, color='green', label='Bon')
            ax.axhspan(50, 100, alpha=0.1, color='yellow', label='Modéré')
            ax.axhspan(100, 300, alpha=0.1, color='red', label='Mauvais')
            
            ax.set_xlabel('Date')
            ax.set_ylabel('IQA')
            ax.set_title(f'Évolution de l\'IQA - {station_name}')
            ax.legend()
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=100)
            buffer.seek(0)
            plt.close()
            
            return buffer
        except Exception as e:
            print(f"Erreur génération graphique: {e}")
            return None

    def generate_excel(self, filename):
        """Génère un rapport Excel professionnel"""
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Données EcoWatch"
        
        # En-têtes
        headers = ['Date/Heure', 'Station', 'IQA', 'PM2.5', 'PM10', 'CO', 'NO2', 'SO2', 'O3', 'Température', 'Humidité']
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for station in self.stations:
            readings = Reading.objects.filter(
                station=station,
                timestamp__date__gte=self.start_date,
                timestamp__date__lte=self.end_date
            ).order_by('timestamp')
            
            for reading in readings:
                ws.append([
                    reading.timestamp.strftime('%d/%m/%Y %H:%M'),
                    station.name,
                    reading.iqa or '',
                    reading.pm25 or '',
                    reading.pm10 or '',
                    reading.co or '',
                    reading.no2 or '',
                    reading.so2 or '',
                    reading.o3 or '',
                    reading.temperature or '',
                    reading.humidity or '',
                ])
        
        # Ajuster les largeurs de colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        wb.save(buffer)
        buffer.seek(0)
        return buffer
