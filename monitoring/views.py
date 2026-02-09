from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib import messages
from django.db.models import Count
from django.db.models.functions import TruncDay
from .forms import CustomUserCreationForm, ContactForm, NewsletterForm, StationForm
from datetime import timedelta, datetime
from rest_framework import viewsets
from django.utils import timezone
from .models import Station, Reading, AlertRule, AlertLog, GeneratedReport
from .serializers import StationSerializer, ReadingSerializer
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, FileResponse, JsonResponse
from django.views.decorators.http import require_POST, require_http_methods
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from .report_generator import ReportGenerator
import os
from django.conf import settings

def index(request):
    stations = Station.objects.all()
    return render(request, 'monitoring/index.html', {'stations': stations})

def dashboard(request):
    return render(request, 'monitoring/dashboard.html')

def contact(request):
    form = ContactForm()
    newsletter_form = NewsletterForm()

    if request.method == 'POST':
        if 'submit_contact' in request.POST:
            form = ContactForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Votre message a été envoyé avec succès !")
                return redirect('contact')
        
        elif 'submit_newsletter' in request.POST:
            newsletter_form = NewsletterForm(request.POST)
            if newsletter_form.is_valid():
                newsletter_form.save()
                messages.success(request, "Merci de vous être inscrit à notre newsletter !")
                return redirect('contact')
            else:
                messages.error(request, "Une erreur est survenue lors de l'inscription à la newsletter.")

    context = {
        'form': form,
        'newsletter_form': newsletter_form,
    }
    return render(request, 'monitoring/contact.html', context)

def features(request):
    return render(request, 'monitoring/features.html')

def about(request):
    return render(request, 'monitoring/about.html')

from django.core.paginator import Paginator

def stations_list(request):
    stations_list = Station.objects.all().order_by('-created_at')
    paginator = Paginator(stations_list, 6) # Show 6 stations per page
    
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'monitoring/stations.html', {
        'page_obj': page_obj,
        'station_form': StationForm()
    })

def add_station(request):
    if request.method == 'POST':
        form = StationForm(request.POST, request.FILES)
        if form.is_valid():
            station = form.save()
            if request.headers.get('HX-Request'):
                messages.success(request, f"La station {station.name} a été ajoutée avec succès !")
                return HttpResponse(status=204, headers={'HX-Trigger': 'stationAdded'})
            
            messages.success(request, f"La station {station.name} a été ajoutée avec succès !")
            return redirect('stations')
    
    return redirect('stations')

def edit_station(request, pk):
    station = get_object_or_404(Station, pk=pk)
    if request.method == 'POST':
        form = StationForm(request.POST, request.FILES, instance=station)
        if form.is_valid():
            form.save()
            if request.headers.get('HX-Request'):
                messages.success(request, f"La station {station.name} a été modifiée avec succès !")
                return HttpResponse(status=204, headers={'HX-Trigger': 'stationUpdated'})
            
            messages.success(request, f"La station {station.name} a été modifiée avec succès !")
            return redirect('stations')
    
    return redirect('stations')

@require_http_methods(['DELETE', 'POST'])
def delete_station(request, pk):
    station = get_object_or_404(Station, pk=pk)
    station_name = station.name
    station.delete()
    
    if request.headers.get('HX-Request'):
        messages.success(request, f"La station {station_name} a été supprimée.")
        return HttpResponse(status=204, headers={'HX-Trigger': 'stationDeleted'})
    
    messages.success(request, f"La station {station_name} a été supprimée.")
    return redirect('stations')

def alerts_view(request):
    # Récupérer la règle d'alerte globale (pour l'exemple)
    rule, created = AlertRule.objects.get_or_create(id=1)
    
    if request.method == 'POST':
        # Mise à jour des seuils
        rule.iqa_threshold = request.POST.get('iqa_threshold') or None
        rule.pm25_threshold = request.POST.get('pm25_threshold') or None
        rule.co_threshold = request.POST.get('co_threshold') or None
        rule.temperature_threshold = request.POST.get('temp_threshold') or None
        
        # Notifications
        rule.email_address = request.POST.get('email')
        rule.phone_number = request.POST.get('phone')
        rule.email_notification = 'email_notif' in request.POST
        rule.sms_notification = 'sms_notif' in request.POST
        
        rule.save()
        messages.success(request, "Configuration des alertes mise à jour.")
        return redirect('alerts')

    time_threshold = timezone.now() - timedelta(hours=24)
    month_threshold = timezone.now() - timedelta(days=30)
    
    # Statistiques réelles
    alerts_all = AlertLog.objects.all().order_by('-created_at')
    alerts_today_count = alerts_all.filter(created_at__gte=time_threshold).count()
    
    # Répartition par station
    alerts_by_station = list(AlertLog.objects.values('station__name').annotate(count=Count('id')).order_by('-count'))
    
    # Filière temporelle (30 derniers jours)
    alerts_by_day = AlertLog.objects.filter(created_at__gte=month_threshold)\
        .annotate(day=TruncDay('created_at'))\
        .values('day')\
        .annotate(count=Count('id'))\
        .order_by('day')
    
    # Formater pour Chart.js
    alerts_by_day_list = [
        {'day': item['day'].strftime('%d/%m'), 'count': item['count']}
        for item in alerts_by_day
    ]

    context = {
        'rule': rule,
        'alerts': alerts_all[:10], # 10 dernières alertes
        'alerts_total': alerts_all.count(),
        'alerts_today': alerts_today_count,
        'rules_count': 1,
        'alerts_by_station': alerts_by_station,
        'alerts_by_day': alerts_by_day_list,
    }
    return render(request, 'monitoring/alerts.html', context)

def reports_view(request):
    stations = Station.objects.all()
    total_readings = Reading.objects.all().count()
    
    # Récupérer l'historique réel des rapports
    generated_reports = GeneratedReport.objects.all()[:10]  # 10 derniers rapports
    total_reports = GeneratedReport.objects.count()
    total_exports = GeneratedReport.objects.filter(created_at__month=timezone.now().month).count()
    
    context = {
        'stations': stations,
        'total_reports': total_reports,
        'total_exports': total_exports,
        'total_readings': total_readings,
        'generated_reports': generated_reports,
    }
    return render(request, 'monitoring/reports.html', context)

def exports_view(request):
    """Centre de téléchargement des rapports et exports archivés"""
    exports = GeneratedReport.objects.all().order_by('-created_at')
    return render(request, 'monitoring/exports.html', {'exports': exports})

def analyses_view(request):
    """Vue pour la page Analyses Avancées 2.0"""
    from django.db.models.functions import TruncMonth, TruncHour
    from django.db.models import Avg, Max, Min
    from django.utils import timezone
    import datetime
    import json

    # 0. PARAMÈTRES DE FILTRAGE
    station_id = request.GET.get('station')
    pollutant = request.GET.get('pollutant', 'iqa')  # défaut: IQA
    
    # Sécurité sur le polluant choisi
    allowed_pollutants = {
        'iqa': 'Indice Qualité Air',
        'pm25': 'PM2.5 (Particules fines)',
        'pm10': 'PM10 (Particules)',
        'co': 'CO (Monoxyde de Carbone)',
        'temperature': 'Température (°C)',
        'humidity': 'Humidité (%)'
    }
    if pollutant not in allowed_pollutants:
        pollutant = 'iqa'

    # Filtre de base
    readings_base = Reading.objects.all()
    if station_id and station_id.isdigit():
        readings_base = readings_base.filter(station_id=station_id)

    # 1. TENDANCES MENSUELLES (Année glissante)
    end_date = timezone.now()
    start_date = end_date - datetime.timedelta(days=365)
    
    stats_monthly = readings_base.filter(
        timestamp__range=(start_date, end_date)
    ).annotate(
        month=TruncMonth('timestamp')
    ).values('month').annotate(
        avg_val=Avg(pollutant)
    ).order_by('month')
    
    labels = []
    data_val = []
    for entry in stats_monthly:
        if entry['month']:
            labels.append(entry['month'].strftime('%b %Y'))
            data_val.append(round(entry['avg_val'] or 0, 1))

    # Fallback si pas de données
    if not labels:
        labels = ['Sep', 'Oct', 'Nov', 'Déc', 'Jan', 'Fév']
        data_val = [40, 45, 50, 60, 55, 48]

    # 2. ANALYSE DE CORRÉLATION (30 derniers jours)
    # On croise le polluant choisi avec la Température et l'Humidité
    start_30d = end_date - datetime.timedelta(days=30)
    correlation_data = readings_base.filter(
        timestamp__range=(start_30d, end_date)
    ).annotate(
        hour=TruncHour('timestamp')
    ).values('hour').annotate(
        val=Avg(pollutant),
        temp=Avg('temperature'),
        hum=Avg('humidity')
    ).order_by('hour')

    corr_labels = []
    corr_vals = []
    corr_temps = []
    corr_hums = []
    
    for entry in correlation_data:
        if entry['hour']:
            corr_labels.append(entry['hour'].strftime('%d/%m %H:00'))
            corr_vals.append(round(entry['val'] or 0, 1))
            corr_temps.append(round(entry['temp'] or 0, 1))
            corr_hums.append(round(entry['hum'] or 0, 1))

    # 3. SMART INSIGHTS (Calcul automatique)
    insight = "Chargement des données..."
    insight_type = "neutral" # neutral, success, warning, danger
    
    # Petit "algo" d'insight
    if data_val:
        current_avg = data_val[-1]
        prev_avg = data_val[-2] if len(data_val) > 1 else current_avg
        diff = current_avg - prev_avg
        
        if diff > 10:
            insight = f"⚠️ Hausse significative de {allowed_pollutants[pollutant]} (+{round(diff,1)}) ce mois-ci par rapport au mois dernier."
            insight_type = "warning"
        elif diff < -10:
            insight = f"✅ Amélioration notable de {allowed_pollutants[pollutant]} (-{round(abs(diff),1)}) constatée ce mois-ci."
            insight_type = "success"
        else:
            insight = f"📊 Stabilité globale de {allowed_pollutants[pollutant]} sur la période récente."
            insight_type = "neutral"

    # 4. HEATMAP (Dernières 24h pour le slider temporal)
    # Pour faire simple, on envoie les readings des dernières 24h avec lat/lng
    start_24h = end_date - datetime.timedelta(days=1)
    readings_24h = Reading.objects.filter(timestamp__range=(start_24h, end_date)).select_related('station')
    
    heatmap_24h = []
    for r in readings_24h:
        if r.station and r.iqa:
            heatmap_24h.append({
                'lat': float(r.station.latitude),
                'lng': float(r.station.longitude),
                'iqa': r.iqa,
                'time': r.timestamp.isoformat()
            })

    # 5. ANALYSE COMPORTEMENTALE (Par heure de la journée)
    from django.db.models.functions import ExtractHour
    
    behavioral_data = readings_base.filter(
        timestamp__range=(start_30d, end_date)
    ).annotate(
        hour_only=ExtractHour('timestamp')
    ).values('hour_only').annotate(
        avg_val=Avg(pollutant)
    ).order_by('hour_only')
    
    behavioral_labels = [f"{h}h" for h in range(24)]
    behavioral_values = [0] * 24
    
    for entry in behavioral_data:
        h = entry['hour_only']
        if 0 <= h < 24:
            behavioral_values[h] = round(entry['avg_val'] or 0, 1)

    # 6. DESCRIPTIONS DYNAMIQUES (Détails Novice/Expert basés sur données RÉELLES)
    current_avg = data_val[-1] if data_val else 0
    prev_avg = data_val[-2] if len(data_val) > 1 else current_avg
    diff = current_avg - prev_avg
    trend_text = "en hausse" if diff > 0 else "en baisse"
    
    all_insights = {
        'iqa': {
            'spatial': f"L'Indice global est de {current_avg}. Une zone <span style='color:#ef4444; font-weight:700;'>Rouge</span> sur cette caméra thermique indique une atmosphère saturée où les polluants stagnent, impactant lourdement la santé.",
            'behavioral': f"Sur un cycle de 24h, l'IQA a varié de {abs(round(diff,1))} points ({trend_text}). Les pics matinaux (7h-9h) révèlent souvent l'impact du trafic alors que les pics nocturnes pointent vers des activités industrielles.",
            'natural': "L'analyse montre comment la météo interagit avec la pollution. De fortes températures sans vent créent souvent un blocage emprisonnant les particules au sol.",
            'long_terme': f"Avec une moyenne de {current_avg} ce mois-ci, ce graphique permet de suivre l'efficacité des politiques de santé. Les pics historiques identifient les cycles critiques comme l'Harmattan.",
            'scale': {'min': 'Pur (Sain)', 'max': 'Saturé (Danger)'}
        },
        'pm25': {
            'spatial': f"Concentration : {current_avg} µg/m³. Les particules fines PM2.5 sont très volatiles. La heatmap localise les foyers de poussières qui s'infiltrent profondément dans les poumons.",
            'behavioral': f"Évolution de {abs(round(diff,1))} µg/m³ ({trend_text}). Ces particules stagnent surtout le matin lors de l'inversion thermique, rendant l'air plus 'lourd' et dangereux à respirer.",
            'natural': "La pluie diminue ces particules par lessivage. À l'inverse, l'air sec et stagnant favorise leur accumulation dangereuse en milieu urbain.",
            'long_terme': f"Moyenne de {current_avg} µg/m³. Crucial pour surveiller l'impact des vents sahariens durant la saison sèche et comparer l'évolution annuelle.",
            'scale': {'min': 'Léger', 'max': 'Critique'}
        },
        'pm10': {
            'spatial': f"Niveau actuel : {current_avg} µg/m³. Les PM10 (plus grosses que les PM2.5) retombent plus vite mais restent critiques près des axes routiers et chantiers.",
            'behavioral': f"Tendance {trend_text} de {abs(round(diff,1))} µg/m³. Les pics correspondent souvent aux activités de construction ou au trafic intense de camions.",
            'natural': "Le vent est le principal facteur de dispersion de ce gaz inodore. Une forte humidité peut les agglomérer et les faire retomber au sol prématurément.",
            'long_terme': "Analyse long terme essentielle pour mesurer l'impact des activités de construction urbaine sur la qualité de l'air locale.",
            'scale': {'min': 'Clair', 'max': 'Poussiéreux'}
        },
        'co': {
            'spatial': f"Taux de CO : {current_avg} ppm. Ce gaz provient de la combustion incomplète. Les zones rouges marquent les carrefours et zones de stockage de déchets.",
            'behavioral': f"Le CO est un traceur direct du trafic : tendance {trend_text} de {abs(round(diff,1))} ppm. Les pics coïncident avec les engorgements routiers majeurs.",
            'natural': "Inodore et toxique, le CO se dissipe difficilement sans courant d'air. Les zones confinées sont les plus à risque lors des pics de chaleur.",
            'long_terme': "Permet de vérifier si le renouvellement des moteurs et l'amélioration de la circulation améliorent la santé respiratoire globale.",
            'scale': {'min': 'Trace', 'max': 'Toxique'}
        },
        'temperature': {
            'spatial': f"Moyenne : {current_avg}°C. La carte identifie les 'îlots de chaleur'. Les zones bétonnées apparaissent plus chaudes que les rares espaces verts urbains.",
            'behavioral': f"Variation de {abs(round(diff,1))}°C ({trend_text}). L'analyse jour/nuit montre la capacité thermique du sol urbain à relâcher la chaleur emmagasinée.",
            'natural': "Une forte chaleur corrélée à une humidité basse augmente le risque de feux et de poussières en suspension dans l'air.",
            'long_terme': "Indicateur direct du réchauffement climatique local et de l'urgence de végétaliser les zones urbaines denses.",
            'scale': {'min': 'Frais', 'max': 'Canicule'}
        },
        'humidity': {
            'spatial': f"Humidité : {current_avg}%. Localise les zones de stagnation humide propices au développement de moisissures ou au maintien des gouttelettes de pollution.",
            'behavioral': f"Tendance {trend_text} de {abs(round(diff,1))}%. L'humidité est maximale la nuit et chute au soleil, influençant la sensation de chaleur (indice de confort).",
            'natural': "Une humidité saturée (proche de 100%) favorise le smog urbain en piégeant les gaz d'échappement dans les gouttelettes d'eau.",
            'long_terme': "Essentiel pour anticiper les saisons de pluies et l'impact de l'humidité sur la pérennité des infrastructures et la santé.",
            'scale': {'min': 'Sec', 'max': 'Humide'}
        }
    }
    
    # Fallback si pollutions exotiques
    current_insights = all_insights.get(pollutant, all_insights['iqa'])

    # 7. TITRES DYNAMIQUES POUR L'EN-TÊTE ET LES INSIGHTS
    insight_title = "Analyse Prédictive & Recommandation"
    if diff > 10:
        insight_title = f"Critique : Dégradation du {allowed_pollutants[pollutant]}"
    elif diff < -10:
        insight_title = f"Optimiste : Amélioration du {allowed_pollutants[pollutant]}"
    
    header_subtitle = f"Analyse précise pour {allowed_pollutants[pollutant]} - Basée sur les stations actives."

    context = {
        'stations': Station.objects.all(),
        'selected_station': station_id,
        'selected_pollutant': pollutant,
        'pollutant_name': allowed_pollutants[pollutant],
        'trend_labels': json.dumps(labels),
        'trend_data': json.dumps(data_val),
        'current_avg_val': current_avg,
        'trend_diff': round(diff, 1),
        'corr_labels': json.dumps(corr_labels),
        'corr_vals': json.dumps(corr_vals),
        'corr_temps': json.dumps(corr_temps),
        'corr_hums': json.dumps(corr_hums),
        'insight': insight,
        'insight_type': insight_type,
        'insight_title': insight_title,
        'header_subtitle': header_subtitle,
        'heatmap_24h': json.dumps(heatmap_24h),
        'behavioral_labels': json.dumps(behavioral_labels),
        'behavioral_values': json.dumps(behavioral_values),
        'pollutant_insights': current_insights,
    }
    
    return render(request, 'monitoring/analyses.html', context)

def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('dashboard')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/signup.html', {'form': form})

class StationViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer

class ReadingViewSet(viewsets.ModelViewSet):
    queryset = Reading.objects.all()
    serializer_class = ReadingSerializer

    def get_queryset(self):
        queryset = Reading.objects.all()
        station_id = self.request.query_params.get('station')
        if station_id:
            queryset = queryset.filter(station_id=station_id)
        return queryset

    def perform_create(self, serializer):
        # Optional: Add logic to calculate IQA if not provided by the sensor
        instance = serializer.save()
        if not instance.iqa:
            # Simple placeholder logic for IQA calculation
            # Based on PM2.5 (very simplified for now)
            instance.iqa = int(instance.pm25 * 2) if instance.pm25 else 0
            instance.save()
        
        # Déclenchement des alertes pour les capteurs physiques également
        from .utils import check_alert_rules
        check_alert_rules(instance)

@require_POST
def generate_report(request):
    """Génère un rapport PDF"""
    try:
        # Récupérer les paramètres
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        station_ids = request.POST.getlist('stations')
        report_type = request.POST.get('report_type', 'CUSTOM')
        
        # Debug logs to file
        with open('debug_pdf.log', 'a', encoding='utf-8') as log:
            log.write(f"\n--- {datetime.now()} ---\n")
            log.write(f"Parameters: start={start_date_str}, end={end_date_str}, stations={station_ids}\n")
        
        if not start_date_str or not end_date_str or not station_ids:
            return JsonResponse({'error': 'Paramètres manquants'}, status=400)

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Récupérer les stations
        stations = Station.objects.filter(id__in=station_ids)
        if not stations.exists():
            return JsonResponse({'error': 'Aucune station trouvée'}, status=400)
        
        # Vérifier les données disponibles
        total_readings = 0
        for station in stations:
            count = Reading.objects.filter(
                station=station,
                timestamp__date__gte=start_date,
                timestamp__date__lte=end_date
            ).count()
            total_readings += count
        
        with open('debug_pdf.log', 'a', encoding='utf-8') as log:
            log.write(f"Total readings found in DB: {total_readings}\n")
            
        if total_readings == 0:
            return JsonResponse({'error': 'Aucune donnée trouvée pour cette période et ces stations'}, status=400)
            
        # Générer le rapport
        report_format = request.POST.get('format', 'PDF')
        generator = ReportGenerator(stations, start_date, end_date, report_type)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if report_format == 'EXCEL':
            filename = f"rapport_ecowatch_{timestamp}.xlsx"
            report_buffer = generator.generate_excel(filename)
        else:
            filename = f"rapport_ecowatch_{timestamp}.pdf"
            report_buffer = generator.generate_pdf(filename)
        
        report_data = report_buffer.getvalue()
        
        with open('debug_pdf.log', 'a', encoding='utf-8') as log:
            log.write(f"Report ({report_format}) generated size: {len(report_data)} bytes\n")
            
        # Sauvegarder le fichier dans media/reports
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, filename)
        
        with open(file_path, 'wb') as f:
            f.write(report_data)
        
        # Créer un enregistrement dans la base de données
        GeneratedReport.objects.create(
            report_type=report_type,
            format=report_format,
            start_date=start_date,
            end_date=end_date,
            file_path=f'reports/{filename}',
            stations_included=[{'id': s.id, 'name': s.name} for s in stations]
        )
        
        # Retourner l'URL du fichier en JSON pour un téléchargement propre côté client
        return JsonResponse({
            'success': True,
            'url': f"{settings.MEDIA_URL}reports/{filename}",
            'filename': filename
        })
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        stack_trace = traceback.format_exc()
        
        with open('debug_pdf.log', 'a', encoding='utf-8') as log:
            log.write(f"❌ CRASH: {error_msg}\n")
            log.write(f"{stack_trace}\n")
            
        print(f"   ❌ ERREUR: {error_msg}")
        return JsonResponse({'error': error_msg}, status=500)

@require_POST
def export_excel(request):
    """Exporte les données vers Excel"""
    try:
        # Paramètres optionnels pour export rapide
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        station_ids = request.POST.getlist('stations')
        
        # Si pas de paramètres, exporter toutes les données récentes
        if not start_date_str or not end_date_str:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
            stations = Station.objects.all()
        else:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            stations = Station.objects.filter(id__in=station_ids) if station_ids else Station.objects.all()
        
        # Créer le workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Données Qualité de l'Air"
        
        # En-têtes
        headers = ['Date/Heure', 'Station', 'IQA', 'PM2.5', 'PM10', 'CO', 'NO2', 'SO2', 'O3', 'Température', 'Humidité']
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for station in stations:
            readings = Reading.objects.filter(
                station=station,
                timestamp__date__gte=start_date,
                timestamp__date__lte=end_date
            ).order_by('timestamp')
            
            for reading in readings:
                ws.append([
                    reading.timestamp.strftime('%d/%m/%Y %H:%M'),
                    station.name,
                    reading.iqa or '',
                    reading.pm25 or '',
                    reading.pm10 or '',
                    reading.co or '',
                    reading.no2 or '',
                    reading.so2 or '',
                    reading.o3 or '',
                    reading.temperature or '',
                    reading.humidity or '',
                ])
        
        # Ajuster les largeurs de colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder dans un buffer
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="ecowatch_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_POST
def export_csv(request):
    """Exporte toutes les données vers CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ecowatch_data_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date/Heure', 'Station', 'Latitude', 'Longitude', 'IQA', 'PM2.5', 'PM10', 'CO', 'NO2', 'SO2', 'O3', 'Température', 'Humidité', 'Source'])
    
    readings = Reading.objects.select_related('station').order_by('-timestamp')[:10000]  # Limiter à 10000 derniers relevés
    
    for reading in readings:
        writer.writerow([
            reading.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            reading.station.name,
            reading.station.latitude,
            reading.station.longitude,
            reading.iqa or '',
            reading.pm25 or '',
            reading.pm10 or '',
            reading.co or '',
            reading.no2 or '',
            reading.so2 or '',
            reading.o3 or '',
            reading.temperature or '',
            reading.humidity or '',
            reading.source_type,
        ])
    
    return response